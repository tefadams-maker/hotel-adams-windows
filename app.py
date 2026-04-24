# ==============================================================================
# 🏨 HOTEL ADAMS - SISTEMA DE GESTIÓN
# Backend: Flask + SQLAlchemy + SQLite
# ==============================================================================

# ==============================================================================
# IMPORTS
# ==============================================================================
from flask import Flask, render_template, request, redirect, url_for, flash, session, send_file, abort
from sqlalchemy import or_, func
from database import db, Habitacion, Huesped, Reserva, Usuario, Producto, CargoExtra
from datetime import datetime, timedelta
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from io import BytesIO
import os

# ==============================================================================
# CONFIGURACIÓN DE LA APLICACIÓN
# ==============================================================================

app = Flask(__name__)
app.config['SECRET_KEY'] = 'hotel-adams-seguro-2025'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///hotel_adams.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db.init_app(app)

# ==============================================================================
# DATOS INICIALES - HABITACIONES Y USUARIOS
# ==============================================================================

HABITACIONES_DATA = [
    # Piso 2do
    ('200', 'S', '2do', 40), ('201', 'Q+M', '2do', 110), ('202', 'M', '2do', 55),
    ('203', 'M+S', '2do', 85), ('204', 'M', '2do', 55), ('205', 'M+S', '2do', 85),
    ('206', 'S', '2do', 45), ('207', 'M', '2do', 55),
    # Piso 3ro
    ('300', 'C', '3ro', 80), ('301', 'J', '3ro', 100), ('302', 'Q+S', '3ro', 95),
    ('303', 'Q+S', '3ro', 95), ('304', 'M', '3ro', 55), ('305', 'Q+S', '3ro', 95),
    ('306', 'M', '3ro', 55), ('307', 'Q+S', '3ro', 95), ('308', 'M', '3ro', 55),
    ('309', 'S', '3ro', 45),
    # Piso 4to
    ('400', 'C', '4to', 80), ('401', 'J', '4to', 100), ('403', 'Q', '4to', 65),
    ('404', 'M', '4to', 55), ('405', 'Q', '4to', 65), ('406', 'M', '4to', 55),
    ('407', 'Q', '4to', 65), ('408', 'M', '4to', 55), ('409', 'S', '4to', 45),
    # Piso 5to
    ('500', 'C', '5to', 80), ('501', 'J', '5to', 100), ('502', 'Q+M', '5to', 110),
    ('503', 'Q+S', '5to', 95), ('504', 'M', '5to', 55), ('505', 'Q+S', '5to', 95),
    ('506', 'M', '5to', 55), ('507', 'S+S', '5to', 80), ('508', 'M', '5to', 55),
    ('509', 'S', '5to', 45),
    # Piso 6to / Especiales
    ('601', 'T', '6to', 120), ('602', 'M', '6to', 55), ('603', 'T', '6to', 120),
    # Departamentos
    ('Dpto-01', 'Q', '6to', 180), ('Dpto-02', 'M+S', '6to', 200),
]

def inicializar_database():
    """Crea tablas y datos iniciales si no existen"""
    with app.app_context():
        db.create_all()
        
        if Habitacion.query.count() == 0:
            for numero, tipo, piso, precio in HABITACIONES_DATA:
                db.session.add(Habitacion(numero=numero, tipo=tipo, piso=piso, precio_base=precio))
            db.session.commit()
        
        if Usuario.query.count() == 0:
            usuarios_seed = [
                ('admin', 'admin@hoteladams.com', 'admin123', 'Administrador General', 'Gerencia'),
                ('recepcionista1', 'recepcion1@hoteladams.com', 'recepcion1', 'María Pérez', 'Recepcion'),
                ('recepcionista2', 'recepcion2@hoteladams.com', 'recepcion2', 'Juan Díaz', 'Recepcion'),
                ('limpieza', 'limpieza@hoteladams.com', 'limpieza1', 'Carlos Mendoza', 'Limpieza'),
            ]
            for username, email, password, nombre, rol in usuarios_seed:
                db.session.add(Usuario(username=username, email=email, password=password, 
                                     nombre_completo=nombre, rol=rol))
            db.session.commit()

# ==============================================================================
# 🔐 AUTENTICACIÓN
# ==============================================================================

@app.route('/login', methods=['GET', 'POST'])
def login():
    if 'user_id' in session:
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        usuario = Usuario.query.filter_by(username=request.form['username']).first()
        
        if usuario and usuario.check_password(request.form['password']) and usuario.activo:
            session.update({
                'user_id': usuario.id,
                'username': usuario.username,
                'usuario_nombre': usuario.nombre_completo,
                'usuario_rol': usuario.rol
            })
            flash('✅ Bienvenido/a!', 'success')
            return redirect(url_for('index'))
        flash('❌ Usuario o contraseña incorrectos', 'danger')
    
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    flash('Has cerrado sesión correctamente', 'info')
    return redirect(url_for('login'))

# ==============================================================================
# 🏠 DASHBOARD PRINCIPAL
# ==============================================================================

@app.route('/')
def index():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    rol_usuario = session.get('usuario_rol', '')
    filtro_piso = request.args.get('piso', '')
    filtro_tipo = request.args.get('tipo', '')
    
    # 🔒 Limpieza solo ve habitaciones Sucias o En Mantenimiento
    if rol_usuario == 'Limpieza':
        habitaciones = Habitacion.query.filter(
            Habitacion.estado_limpieza.in_(['Sucia', 'En Mantenimiento'])
        ).order_by(Habitacion.piso, Habitacion.numero).all()
    else:
        habitaciones = Habitacion.query.order_by(Habitacion.piso, Habitacion.numero).all()
    
    # Aplicar filtros de búsqueda
    if filtro_piso:
        habitaciones = [h for h in habitaciones if h.piso == filtro_piso]
    if filtro_tipo:
        habitaciones = [h for h in habitaciones if h.tipo.lower() == filtro_tipo.lower()]
    
    reservas_activas = Reserva.query.filter_by(estado='Activa').all()
    hoy = datetime.now().date()
    salen_hoy = [r for r in reservas_activas if r.fecha_salida and r.fecha_salida.date() == hoy]
    
    ocupacion = {
        r.habitacion.numero: {
            'huesped': r.huesped.nombre,
            'fecha_entrada': r.fecha_entrada.strftime('%d/%m %H:%M'),
            'fecha_salida': r.fecha_salida.strftime('%d/%m') if r.fecha_salida else '-',
            'id_reserva': r.id
        } for r in reservas_activas
    }
    
    disponibles = len([h for h in habitaciones if h.estado_ocupacion == 'Disponible'])
    ocupadas = len([h for h in habitaciones if h.estado_ocupacion == 'Ocupada'])
    sucias = len([h for h in habitaciones if h.estado_limpieza == 'Sucia'])
    
    return render_template('index.html', 
        habitaciones=habitaciones, 
        ocupacion=ocupacion, 
        salen_hoy=salen_hoy, 
        disponibles=disponibles, 
        ocupadas=ocupadas, 
        sucias=sucias, 
        pisos=['2do', '3ro', '4to', '5to', '6to'],
        tipos=['S', 'M', 'Q', 'J', 'C', 'T', 'Q+M', 'Q+S', 'M+S', 'S+S'],
        usuario=session.get('usuario_nombre'),
        rol=session.get('usuario_rol'),
        filtro_piso=filtro_piso, 
        filtro_tipo=filtro_tipo,
        rol_usuario=rol_usuario)

# ==============================================================================
# 📋 CHECK-IN / RESERVAS (CON FIX PARA ?hab=ID)
# ==============================================================================

@app.route('/checkin', methods=['GET', 'POST'])
def checkin():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    # 🔧 FIX: Inicializar variable para evitar UnboundLocalError
    hab_id_seleccionada = None
    
    # 🔥 Leer habitación seleccionada desde el Dashboard (?hab=123)
    if request.method == 'GET':
        hab_id_seleccionada = request.args.get('hab', type=int)
    
    today_str = datetime.now().strftime('%Y-%m-%d')
    today_date = datetime.now().date()
    
    if request.method == 'POST':
        habitacion = Habitacion.query.get(request.form['habitacion_id'])
        fecha_entrada = datetime.strptime(request.form['fecha_entrada'], '%Y-%m-%d')
        fecha_salida = datetime.strptime(request.form['fecha_salida'], '%Y-%m-%d')
        
        if fecha_salida <= fecha_entrada:
            flash('❌ La fecha de salida debe ser posterior a la de entrada', 'danger')
            return redirect(url_for('checkin'))
        
        # Validar conflictos de fechas
        for r in Reserva.query.filter_by(habitacion_id=habitacion.id, estado='Activa').all():
            if fecha_entrada < r.fecha_salida and fecha_salida > r.fecha_entrada:
                flash('❌ Esta habitación ya tiene una reserva para esas fechas', 'danger')
                return redirect(url_for('checkin'))
        
        # Huésped
        huesped = Huesped.query.filter_by(dni=request.form['dni']).first()
        if not huesped:
            huesped = Huesped(
                nombre=request.form['nombre'],
                apellido=request.form.get('apellido', ''),
                dni=request.form['dni'],
                celular=request.form.get('celular', ''),
                email=request.form.get('email', ''),
                nacionalidad=request.form.get('nacionalidad', 'Perú')
            )
            db.session.add(huesped)
            db.session.commit()
        
        # Reserva
        dias = max(1, (fecha_salida - fecha_entrada).days)
        reserva = Reserva(
            habitacion_id=habitacion.id,
            huesped_id=huesped.id,
            fecha_entrada=fecha_entrada,
            fecha_salida=fecha_salida,
            precio_total=habitacion.precio_base * dias,
            precio_pagado=float(request.form.get('pago', 0)),
            metodo_pago=request.form.get('metodo_pago', 'Efectivo'),
            observaciones=request.form.get('observaciones', ''),
            estado='Activa'
        )
        db.session.add(reserva)
        
        # Estado de habitación
        habitacion.estado_ocupacion = 'Ocupada'
        if fecha_entrada.date() == today_date:
            habitacion.estado_limpieza = 'Sucia'
        
        db.session.commit()
        flash(f'✅ Reserva confirmada para el {request.form["fecha_entrada"]}', 'success')
        return redirect(url_for('index'))
    
        # GET: Mostrar formulario
    # 🔥 Ya no filtramos por estado, mostramos TODAS las habitaciones
    habitaciones = Habitacion.query.order_by(Habitacion.piso, Habitacion.numero).all()
    
    # Traemos las reservas activas para calcular disponibilidad real por fechas
    reservas_activas = Reserva.query.filter_by(estado='Activa').all()
    
    return render_template('checkin.html', 
                          habitaciones=habitaciones, 
                          hoy=today_str,
                          reservas_activas=reservas_activas, 
                          hab_id_seleccionada=hab_id_seleccionada)

# ==============================================================================
# 🚪 CHECK-OUT
# ==============================================================================
@app.route('/checkout/<int:reserva_id>', methods=['GET', 'POST'])
def checkout(reserva_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    reserva = Reserva.query.get_or_404(reserva_id)
    hoy_str = datetime.now().strftime('%Y-%m-%d')
    
    if request.method == 'POST':
        # 1. Obtener fecha real de salida
        fecha_real_str = request.form.get('fecha_real_salida')
        # Si no hay fecha, usa hoy
        fecha_real = datetime.strptime(fecha_real_str, '%Y-%m-%d') if fecha_real_str else datetime.now()
        
        # 2. Calcular extras (Productos)
        total_extras = sum(c.cantidad * c.precio_unitario for c in reserva.cargos_extras)

        # 3. Recalcular días y precio de habitación
        # max(1, ...) asegura que cobre al menos 1 día si las fechas son iguales
        dias_reales = max(1, (fecha_real - reserva.fecha_entrada).days)
        precio_habitacion = reserva.habitacion.precio_base * dias_reales
        
        # El precio final es Habitación + Extras
        precio_final = precio_habitacion + total_extras
        
        # 4. Calcular diferencia (Pago - Total)
        diferencia = reserva.precio_pagado - precio_final
        
        # 🔧 FIX: Manejar error si el campo viene vacío
        pago_adicional_str = request.form.get('pago_adicional', '0')
        pago_adicional = float(pago_adicional_str) if pago_adicional_str else 0.0
        
        reserva.precio_pagado += pago_adicional
        
        if pago_adicional > 0:
            reserva.metodo_pago = request.form.get('metodo_pago_saldo', 'Efectivo')
        
        # 5. Guardar cambios
        reserva.fecha_salida = fecha_real
        reserva.precio_total = precio_final  # Guardamos el total real (Hab + Extras)
        reserva.estado = 'Finalizada'
        reserva.habitacion.estado_ocupacion = 'Disponible'
        reserva.habitacion.estado_limpieza = 'Sucia'
        
        db.session.commit()
        
        # Mensajes
        if diferencia > 0.5:
            flash(f'✅ Check-out realizado. ⚠️ FALTA DEVOLVER S/ {diferencia:.2f} al huésped.', 'warning')
        elif diferencia < -0.5:
            flash(f'⚠️ Check-out realizado. El huésped aún debe S/ {abs(diferencia):.2f}.', 'danger')
        else:
            flash('✅ Check-out realizado. Cuenta saldada perfectamente.', 'success')
            
        return redirect(url_for('index'))
    
    # GET: Calcular totales para mostrar en pantalla
    total_extras = sum(c.cantidad * c.precio_unitario for c in reserva.cargos_extras)
    
    return render_template('checkout.html', 
                          reserva=reserva, 
                          hoy=hoy_str, 
                          total_extras=total_extras)

# ==============================================================================
# 🧹 LIMPIEZA
# ==============================================================================

@app.route('/limpieza/<int:habitacion_id>/<estado>')
def limpieza(habitacion_id, estado):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    habitacion = Habitacion.query.get_or_404(habitacion_id)
    habitacion.estado_limpieza = estado
    db.session.commit()
    flash(f'✅ Estado actualizado a: {estado}', 'success')
    return redirect(url_for('index'))

# ==============================================================================
# 📜 HISTORIAL Y BÚSQUEDA
# ==============================================================================

@app.route('/historial')
def historial():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    fecha_inicio = request.args.get('fecha_inicio')
    fecha_fin = request.args.get('fecha_fin')
    estado = request.args.get('estado')
    busqueda = request.args.get('busqueda')
    habitacion_filtro = request.args.get('habitacion')
    
    query = Reserva.query
    
    if fecha_inicio:
        query = query.filter(Reserva.fecha_entrada >= datetime.strptime(fecha_inicio, '%Y-%m-%d'))
    if fecha_fin:
        query = query.filter(Reserva.fecha_entrada <= datetime.strptime(fecha_fin, '%Y-%m-%d'))
    if estado:
        query = query.filter(Reserva.estado == estado)
    
    # 🔥 CORRECCIÓN AQUÍ: Debe estar al mismo nivel que los demás 'if'
    if habitacion_filtro:
        query = query.join(Habitacion).filter(Habitacion.numero.contains(habitacion_filtro))
        
    if busqueda:
        query = query.join(Huesped).filter(
            or_(
                Huesped.nombre.contains(busqueda),
                Huesped.apellido.contains(busqueda),
                Huesped.dni.contains(busqueda)
            )
        )
    
    reservas = query.order_by(Reserva.fecha_entrada.desc()).all()
    return render_template('historial.html', reservas=reservas)

@app.route('/buscar')
def buscar():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    query = request.args.get('q', '')
    resultados = []
    if query:
        for h in Huesped.query.filter((Huesped.nombre.contains(query)) | (Huesped.dni.contains(query))).all():
            resultados.extend(h.reservas)
    return render_template('buscar.html', resultados=resultados, query=query)

# ==============================================================================
# 📅 CALENDARIO
# ==============================================================================

@app.route('/calendario')
def calendario():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    habitaciones = Habitacion.query.order_by(Habitacion.numero).all()
    hoy = datetime.now()
    dias = [(hoy + timedelta(days=i)).strftime('%d/%m') for i in range(30)]
    reservas = Reserva.query.filter_by(estado='Activa').all()
    
    ocupacion = {}
    for hab in habitaciones:
        ocupacion[hab.numero] = []
        for i in range(30):
            fecha = hoy + timedelta(days=i)
            ocupada = any(
                r.habitacion.numero == hab.numero and 
                r.fecha_entrada.date() <= fecha.date() and 
                (not r.fecha_salida or r.fecha_salida.date() >= fecha.date())
                for r in reservas
            )
            if ocupada:
                ocupacion[hab.numero].append('ocupado')
            elif hab.estado_limpieza == 'Sucia':
                ocupacion[hab.numero].append('sucia')
            else:
                ocupacion[hab.numero].append('disponible')
    
    return render_template('calendario.html', habitaciones=habitaciones, dias=dias, ocupacion=ocupacion)

# ==============================================================================
# ⚙️ PANEL DE ADMINISTRACIÓN
# ==============================================================================

@app.route('/admin')
def admin():
    if 'user_id' not in session or session.get('usuario_rol') not in ['Gerencia', 'Recepcion']:
        flash('❌ No tienes permiso para acceder', 'danger')
        return redirect(url_for('index'))
    return render_template('admin.html', 
        habitaciones=Habitacion.query.order_by(Habitacion.numero).all(),
        usuarios=Usuario.query.all())

@app.route('/crear_usuario', methods=['GET', 'POST'])
def crear_usuario():
    if session.get('usuario_rol') != 'Gerencia':
        flash('❌ Solo Gerencia puede crear usuarios', 'danger')
        return redirect(url_for('admin'))
    
    if request.method == 'POST':
        if Usuario.query.filter_by(username=request.form['username']).first():
            flash('⚠️ Este usuario ya existe', 'warning')
            return redirect(url_for('admin'))
        
        db.session.add(Usuario(
            username=request.form['username'],
            email=request.form['email'],
            password=request.form['password'],
            nombre_completo=request.form['nombre_completo'],
            rol=request.form['rol']
        ))
        db.session.commit()
        flash('✅ Usuario creado exitosamente', 'success')
    return redirect(url_for('admin'))

@app.route('/eliminar_usuario/<int:user_id>', methods=['GET', 'POST'])
def eliminar_usuario(user_id):
    if session.get('usuario_rol') != 'Gerencia' or user_id == session['user_id']:
        flash('⚠️ No puedes eliminar este usuario', 'warning')
        return redirect(url_for('admin'))
    
    db.session.delete(Usuario.query.get_or_404(user_id))
    db.session.commit()
    flash('✅ Usuario eliminado correctamente', 'success')
    return redirect(url_for('admin'))

@app.route('/actualizar_precio/<int:hab_id>', methods=['GET', 'POST'])
def actualizar_precio(hab_id):
    if session.get('usuario_rol') not in ['Gerencia', 'Recepcion']:
        flash('❌ No tienes permiso', 'danger')
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        habitacion = Habitacion.query.get_or_404(hab_id)
        habitacion.precio_base = float(request.form['precio_nuevo'])
        db.session.commit()
        flash(f'✅ Precio actualizado a S/ {request.form["precio_nuevo"]}', 'success')
    return redirect(url_for('admin'))

@app.route('/actualizar_tipo/<int:hab_id>', methods=['POST'])
def actualizar_tipo(hab_id):
    if session.get('usuario_rol') not in ['Gerencia', 'Recepcion']:
        flash('❌ No tienes permiso', 'danger')
        return redirect(url_for('index'))
    
    habitacion = Habitacion.query.get_or_404(hab_id)
    habitacion.tipo = request.form['nuevo_tipo']
    db.session.commit()
    flash(f'✅ Tipo actualizado a: {request.form["nuevo_tipo"]}', 'success')
    return redirect(url_for('admin'))

# ==============================================================================
# 📊 EXPORTAR REPORTES A EXCEL
# ==============================================================================

def _generar_excel_reservas(reservas, titulo, nombre_archivo):
    """Función auxiliar para generar Excel"""
    wb, ws = Workbook(), Workbook().active
    ws.title = titulo
    
    headers = ['ID', 'Habitación', 'Huésped', 'Entrada', 'Salida', 'Total S/', 'Pagado S/', 'Método', 'Estado']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h).font = Font(bold=True)
    
    for row, r in enumerate(reservas, 2):
        ws.cell(row=row, column=1, value=r.id)
        ws.cell(row=row, column=2, value=r.habitacion.numero)
        ws.cell(row=row, column=3, value=f"{r.huesped.nombre} {r.huesped.apellido}")
        ws.cell(row=row, column=4, value=r.fecha_entrada.strftime('%d/%m %H:%M'))
        ws.cell(row=row, column=5, value=r.fecha_salida.strftime('%d/%m %H:%M') if r.fecha_salida else '-')
        ws.cell(row=row, column=6, value=f"S/ {r.precio_total:.2f}")
        ws.cell(row=row, column=7, value=f"S/ {r.precio_pagado:.2f}")
        ws.cell(row=row, column=8, value=r.metodo_pago)
        ws.cell(row=row, column=9, value=r.estado)
    
    for col in range(1, 10):
        ws.column_dimensions[chr(col + 64)].width = 15
    
    path = os.path.join('reportes', f'{nombre_archivo}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
    os.makedirs('reportes', exist_ok=True)
    wb.save(path)
    return path

@app.route('/export/reservas')
def export_reservas_excel():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    path = _generar_excel_reservas(
        Reserva.query.order_by(Reserva.fecha_entrada.desc()).all(),
        'Reservas - Hotel Adams', 'Reservas'
    )
    flash(f'✅ Reporte generado: {os.path.basename(path)}', 'success')
    return redirect(url_for('historial'))

@app.route('/export/huespedes')
def export_huespedes_excel():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    wb, ws = Workbook(), Workbook().active
    ws.title = "Huéspedes - Hotel Adams"
    
    headers = ['ID', 'Nombre', 'Apellido', 'DNI', 'Celular', 'Email', 'Nacionalidad']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h).font = Font(bold=True)
    
    for row, h in enumerate(Huesped.query.all(), 2):
        for col, val in enumerate([h.id, h.nombre, h.apellido, h.dni, h.celular or '-', h.email or '-', h.nacionalidad], 1):
            ws.cell(row=row, column=col, value=val)
    
    path = os.path.join('reportes', f'Huespedes_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
    os.makedirs('reportes', exist_ok=True)
    wb.save(path)
    flash(f'✅ Reporte generado: {os.path.basename(path)}', 'success')
    return redirect(url_for('buscar'))

@app.route('/export/ingresos', methods=['GET', 'POST'])
def export_ingresos_excel():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    fecha_inicio, fecha_fin = None, None
    if request.form.get('fecha_inicio'):
        try: fecha_inicio = datetime.strptime(request.form['fecha_inicio'], '%Y-%m-%d').date()
        except: pass
    if request.form.get('fecha_fin'):
        try: fecha_fin = datetime.strptime(request.form['fecha_fin'], '%Y-%m-%d').date()
        except: pass
    
    query = Reserva.query.filter_by(estado='Finalizada')
    if fecha_inicio: query = query.filter(Reserva.fecha_entrada >= fecha_inicio)
    if fecha_fin: query = query.filter(Reserva.fecha_entrada <= fecha_fin)
    reservas = query.order_by(Reserva.fecha_entrada.desc()).all()
    
    wb, ws = Workbook(), Workbook().active
    ws.title = "Ingresos - Hotel Adams"
    
    ws['A1'], ws['A2'], ws['A3'] = "--- REPORTE DE INGRESOS ---", f"Desde: {fecha_inicio or 'Todo'}", f"Hasta: {fecha_fin or 'Hoy'}"
    ws.merge_cells('A1:H1'); ws['A1'].font = Font(bold=True)
    
    headers = ['Fecha', 'Habitación', 'Huésped', 'Días', 'Método', 'Total', 'Pagado', 'Saldo']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=h)
        cell.font, cell.alignment = Font(bold=True), Alignment(horizontal='center')
    
    total_general, total_pagado = 0, 0
    for row, r in enumerate(reservas, 6):
        dias = 1 if not r.fecha_salida else (r.fecha_salida - r.fecha_entrada).days + 1
        saldo = r.precio_total - r.precio_pagado
        for col, val in enumerate([
            r.fecha_entrada.strftime('%d/%m/%Y'), r.habitacion.numero,
            f"{r.huesped.nombre} {r.huesped.apellido}", dias, r.metodo_pago,
            f"S/ {r.precio_total:.2f}", f"S/ {r.precio_pagado:.2f}", f"S/ {saldo:.2f}"
        ], 1):
            ws.cell(row=row, column=col, value=val)
        total_general += r.precio_total
        total_pagado += r.precio_pagado
    
    ws['J5'] = "--- TOTALES POR MÉTODO ---"; ws.merge_cells('J5:L5'); ws['J5'].font = Font(bold=True)
    ws['J7'], ws['K7'], ws['L7'] = "Método", "Total", "%"
    
    totales = {}
    for r in reservas:
        metodo = r.metodo_pago or 'No especificado'
        totales[metodo] = totales.get(metodo, 0) + r.precio_pagado
    
    for row_tot, (metodo, monto) in enumerate(sorted(totales.items()), 8):
        pct = (monto / total_general * 100) if total_general > 0 else 0
        ws.cell(row=row_tot, column=10, value=metodo)
        ws.cell(row=row_tot, column=11, value=f"S/ {monto:.2f}")
        ws.cell(row=row_tot, column=12, value=f"{pct:.1f}%")
    
    ws.cell(row=row_tot+2, column=12, value=f"S/ {total_pagado:.2f}")
    ws.cell(row=row_tot+4, column=12, value=f"S/ {total_general - total_pagado:.2f}")
    
    for col in range(1, 9):
        max_len = max((len(str(cell.value)) for cell in ws[chr(col + 64)] if cell.value), default=0)
        ws.column_dimensions[chr(col + 64)].width = max_len + 2
    
    path = os.path.join('reportes', f'Finanzas_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
    os.makedirs('reportes', exist_ok=True)
    wb.save(path)
    flash(f'✅ Reporte financiero generado: {os.path.basename(path)}', 'success')
    return redirect(url_for('index'))

# ==============================================================================
# 📂 FICHA DEL HUÉSPED Y PANEL LIMPIEZA
# ==============================================================================

@app.route('/ficha/<int:huesped_id>')
def ficha_huesped(huesped_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    huesped = Huesped.query.get_or_404(huesped_id)
    reservas = Reserva.query.filter_by(huesped_id=huesped_id).order_by(Reserva.fecha_entrada.desc()).all()
    
    return render_template('ficha_huesped.html', huesped=huesped, reservas=reservas)

@app.route('/panel-limpieza')
def panel_limpieza():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    sucias = Habitacion.query.filter_by(estado_limpieza='Sucia').order_by(Habitacion.piso, Habitacion.numero).all()
    limpias = Habitacion.query.filter_by(estado_limpieza='Limpia').order_by(Habitacion.piso, Habitacion.numero).all()
    
    return render_template('panel_limpieza.html', sucias=sucias, limpias=limpias)

# ==============================================================================
# 🛒 INVENTARIO Y CARGO DE EXTRAS
# ==============================================================================

@app.route('/productos')
def productos():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    prods = Producto.query.order_by(Producto.nombre).all()
    return render_template('productos.html', productos=prods)

@app.route('/productos/crear', methods=['POST'])
def crear_producto():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    db.session.add(Producto(
        nombre=request.form['nombre'],
        precio=float(request.form['precio']),
        stock=int(request.form['stock']),
        categoria=request.form.get('categoria', 'General')
    ))
    db.session.commit()
    flash('✅ Producto agregado al inventario', 'success')
    return redirect(url_for('productos'))

@app.route('/productos/editar/<int:prod_id>', methods=['POST'])
def editar_producto(prod_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    prod = Producto.query.get_or_404(prod_id)
    prod.nombre = request.form['nombre']
    prod.precio = float(request.form['precio'])
    prod.stock = int(request.form['stock'])
    prod.categoria = request.form.get('categoria', 'General')
    db.session.commit()
    flash('✅ Producto actualizado', 'success')
    return redirect(url_for('productos'))

@app.route('/productos/eliminar/<int:prod_id>')
def eliminar_producto(prod_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    db.session.delete(Producto.query.get_or_404(prod_id))
    db.session.commit()
    flash('🗑️ Producto eliminado', 'info')
    return redirect(url_for('productos'))

@app.route('/cargo-extra/<int:reserva_id>', methods=['GET', 'POST'])
def cargo_extra(reserva_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    reserva = Reserva.query.get_or_404(reserva_id)
    if reserva.estado != 'Activa':
        flash('❌ No se pueden agregar extras a reservas finalizadas', 'danger')
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        prod_id = int(request.form['producto_id'])
        cantidad = int(request.form['cantidad'])
        
        prod = Producto.query.get(prod_id)
        if not prod or prod.stock < cantidad:
            flash('❌ Stock insuficiente', 'danger')
            return redirect(url_for('cargo_extra', reserva_id=reserva_id))
        
        cargo = CargoExtra(
            reserva_id=reserva_id,
            producto_id=prod_id,
            cantidad=cantidad,
            precio_unitario=prod.precio
        )
        db.session.add(cargo)
        prod.stock -= cantidad
        db.session.commit()
        
        total_cargo = cantidad * prod.precio
        flash(f'✅ Se cargó {cantidad}x {prod.nombre} (+S/ {total_cargo:.2f}) a la cuenta', 'success')
        return redirect(url_for('checkout', reserva_id=reserva_id))
    
    prods_disponibles = Producto.query.filter(Producto.stock > 0).all()
    return render_template('cargo_extra.html', reserva=reserva, productos=prods_disponibles)

# ==============================================================================
# 🚀 INICIO DE LA APLICACIÓN
# ==============================================================================
# ==============================================================================
# 📊 REPORTES WEB + EXPORTACIÓN EXCEL (VERSIÓN ADMINISTRADORA ✅)
# ==============================================================================

@app.route('/reportes')
def reportes():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    if session.get('usuario_rol') != 'Gerencia':
        flash('❌ Solo Gerencia puede ver reportes', 'danger')
        return redirect(url_for('index'))
    
    # Filtros
    fecha_inicio = request.args.get('fecha_inicio', datetime.now().strftime('%Y-%m-%d'))
    fecha_fin = request.args.get('fecha_fin', datetime.now().strftime('%Y-%m-%d'))
    solo_hoy = request.args.get('solo_hoy', 'false') == 'true'
    
    if solo_hoy:
        fecha_inicio = fecha_fin = datetime.now().strftime('%Y-%m-%d')
    
    try:
        f_ini = datetime.strptime(fecha_inicio, '%Y-%m-%d')
        f_fin = datetime.strptime(fecha_fin, '%Y-%m-%d')
    except:
        f_ini = f_fin = datetime.now()

    # 💰 RESUMEN POR MÉTODO DE PAGO (Lo más importante para caja)
    resumen_metodos = db.session.query(
        Reserva.metodo_pago,
        func.count(Reserva.id).label('cantidad'),
        func.sum(Reserva.precio_pagado).label('total_pagado'),
        func.sum(Reserva.precio_total).label('total_facturado')
    ).filter(
        Reserva.estado == 'Finalizada',
        Reserva.fecha_salida.between(f_ini, f_fin)
    ).group_by(Reserva.metodo_pago).all()

    # 🧾 DETALLE DE CHECK-OUTS (Para conciliar con vouchers)
    detalle_checkouts = db.session.query(
        Reserva.id,
        Reserva.fecha_salida,
        Habitacion.numero.label('habitacion'),
        Huesped.nombre,
        Huesped.apellido,
        Reserva.precio_total,
        Reserva.precio_pagado,
        Reserva.metodo_pago,
        (Reserva.precio_total - Reserva.precio_pagado).label('pendiente')
    ).join(Habitacion).join(Huesped)\
     .filter(Reserva.estado == 'Finalizada', Reserva.fecha_salida.between(f_ini, f_fin))\
     .order_by(Reserva.fecha_salida.desc()).all()

    # 🍹 CONSUMO MINI-BAR (Para reposición de stock)
    consumo_bar = db.session.query(
        Producto.nombre,
        Producto.categoria,
        func.sum(CargoExtra.cantidad).label('unidades'),
        func.sum(CargoExtra.cantidad * CargoExtra.precio_unitario).label('recaudacion')
    ).join(CargoExtra, CargoExtra.producto_id == Producto.id)\
     .join(Reserva, Reserva.id == CargoExtra.reserva_id)\
     .filter(Reserva.estado == 'Finalizada', Reserva.fecha_salida.between(f_ini, f_fin))\
     .group_by(Producto.nombre, Producto.categoria)\
     .order_by(func.sum(CargoExtra.cantidad * CargoExtra.precio_unitario).desc()).all()

    # ⚠️ DEUDAS PENDIENTES (Para gestión de cobranza)
    deudas_pendientes = db.session.query(
        Reserva.id,
        Habitacion.numero.label('habitacion'),
        Huesped.nombre,
        Huesped.apellido,
        Huesped.celular,
        (Reserva.precio_total - Reserva.precio_pagado).label('monto_pendiente'),
        Reserva.fecha_salida
    ).join(Habitacion).join(Huesped)\
     .filter(Reserva.estado == 'Finalizada', 
             Reserva.precio_total > Reserva.precio_pagado,
             Reserva.fecha_salida.between(f_ini, f_fin))\
     .order_by((Reserva.precio_total - Reserva.precio_pagado).desc()).all()

    # 📊 TOTALES GENERALES
    total_facturado = sum(r.total_facturado or 0 for r in resumen_metodos)
    total_pagado = sum(r.total_pagado or 0 for r in resumen_metodos)
    total_pendiente = total_facturado - total_pagado
    total_bar = sum(c.recaudacion or 0 for c in consumo_bar)

    return render_template('reportes.html',
                           fecha_inicio=fecha_inicio,
                           fecha_fin=fecha_fin,
                           solo_hoy=solo_hoy,
                           resumen_metodos=resumen_metodos,
                           detalle_checkouts=detalle_checkouts,
                           consumo_bar=consumo_bar,
                           deudas_pendientes=deudas_pendientes,
                           total_facturado=total_facturado,
                           total_pagado=total_pagado,
                           total_pendiente=total_pendiente,
                           total_bar=total_bar)


@app.route('/export/reportes-excel')
def export_reportes_excel():
    if 'user_id' not in session or session.get('usuario_rol') != 'Gerencia':
        abort(403)

    fecha_inicio = request.args.get('fecha_inicio')
    fecha_fin = request.args.get('fecha_fin')
    f_ini = datetime.strptime(fecha_inicio, '%Y-%m-%d') if fecha_inicio else datetime.now()
    f_fin = datetime.strptime(fecha_fin, '%Y-%m-%d') if fecha_fin else datetime.now()

    # Mismas consultas que en la vista web
    resumen_metodos = db.session.query(
        Reserva.metodo_pago,
        func.count(Reserva.id),
        func.sum(Reserva.precio_pagado),
        func.sum(Reserva.precio_total)
    ).filter(Reserva.estado == 'Finalizada', Reserva.fecha_salida.between(f_ini, f_fin)
    ).group_by(Reserva.metodo_pago).all()

    detalle_checkouts = db.session.query(
        Reserva.fecha_salida, Habitacion.numero, 
        Huesped.nombre, Huesped.apellido,
        Reserva.precio_total, Reserva.precio_pagado,
        (Reserva.precio_total - Reserva.precio_pagado),
        Reserva.metodo_pago
    ).join(Habitacion).join(Huesped)\
     .filter(Reserva.estado == 'Finalizada', Reserva.fecha_salida.between(f_ini, f_fin))\
     .order_by(Reserva.fecha_salida.desc()).all()

    consumo_bar = db.session.query(
        Producto.nombre, Producto.categoria,
        func.sum(CargoExtra.cantidad),
        func.sum(CargoExtra.cantidad * CargoExtra.precio_unitario)
    ).join(CargoExtra, CargoExtra.producto_id == Producto.id)\
     .join(Reserva, Reserva.id == CargoExtra.reserva_id)\
     .filter(Reserva.estado == 'Finalizada', Reserva.fecha_salida.between(f_ini, f_fin))\
     .group_by(Producto.nombre, Producto.categoria).all()

    deudas = db.session.query(
        Habitacion.numero, Huesped.nombre, Huesped.apellido,
        (Reserva.precio_total - Reserva.precio_pagado), Reserva.fecha_salida
    ).join(Habitacion).join(Huesped)\
     .filter(Reserva.estado == 'Finalizada', Reserva.precio_total > Reserva.precio_pagado,
             Reserva.fecha_salida.between(f_ini, f_fin)).all()

    # 📄 CREAR EXCEL CON 4 HOJAS + FÓRMULAS
    wb = Workbook()
    wb.remove(wb.active)  # Eliminar hoja por defecto

    # === HOJA 1: RESUMEN EJECUTIVO ===
    ws_resumen = wb.create_sheet("📊 RESUMEN")
    ws_resumen['A1'] = "REPORTE DE CAJA - HOTEL ADAMS"
    ws_resumen['A1'].font = Font(bold=True, size=14)
    ws_resumen.merge_cells('A1:D1')
    
    ws_resumen['A3'] = "Periodo:"
    ws_resumen['B3'] = f"{f_ini.strftime('%d/%m/%Y')} al {f_fin.strftime('%d/%m/%Y')}"
    ws_resumen['A4'] = "Generado:"
    ws_resumen['B4'] = datetime.now().strftime('%d/%m/%Y %H:%M')
    
    # Totales destacados
    total_facturado = sum(r[3] or 0 for r in resumen_metodos)
    total_pagado = sum(r[2] or 0 for r in resumen_metodos)
    
    ws_resumen['A6'] = "TOTAL FACTURADO"
    ws_resumen['B6'] = total_facturado
    ws_resumen['B6'].number_format = '"S/ "#,##0.00'
    ws_resumen['B6'].font = Font(bold=True, size=12, color="0066CC")
    
    ws_resumen['A7'] = "TOTAL COBRADO"
    ws_resumen['B7'] = total_pagado
    ws_resumen['B7'].number_format = '"S/ "#,##0.00'
    ws_resumen['B7'].font = Font(bold=True, size=12, color="009900")
    
    ws_resumen['A8'] = "PENDIENTE DE COBRO"
    ws_resumen['B8'] = total_facturado - total_pagado
    ws_resumen['B8'].number_format = '"S/ "#,##0.00'
    ws_resumen['B8'].font = Font(bold=True, size=12, color="CC0000")
    
    # Desglose por método
    ws_resumen['A10'] = "DESGLOSE POR MÉTODO DE PAGO"
    ws_resumen['A10'].font = Font(bold=True)
    headers_metodo = ["Método", "Transacciones", "Cobrado", "Facturado"]
    for c, h in enumerate(headers_metodo, 1):
        ws_resumen.cell(row=11, column=c, value=h).font = Font(bold=True)
        ws_resumen.cell(row=11, column=c).fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    for r_idx, (metodo, cant, cobrado, facturado) in enumerate(resumen_metodos, 12):
        ws_resumen.cell(row=r_idx, column=1, value=metodo or "No especificado")
        ws_resumen.cell(row=r_idx, column=2, value=cant or 0)
        cell_cobrado = ws_resumen.cell(row=r_idx, column=3, value=cobrado or 0)
        cell_cobrado.number_format = '"S/ "#,##0.00'
        cell_fact = ws_resumen.cell(row=r_idx, column=4, value=facturado or 0)
        cell_fact.number_format = '"S/ "#,##0.00'
    
    # Fila de totales con FÓRMULA de Excel (no valor fijo)
    fila_total = len(resumen_metodos) + 13
    ws_resumen.cell(row=fila_total, column=1, value="TOTAL").font = Font(bold=True)
    ws_resumen.cell(row=fila_total, column=3, value=f"=SUM(C12:C{fila_total-1})").number_format = '"S/ "#,##0.00'
    ws_resumen.cell(row=fila_total, column=4, value=f"=SUM(D12:D{fila_total-1})").number_format = '"S/ "#,##0.00'
    
    for col in 'ABCD':
        ws_resumen.column_dimensions[col].width = 18

    # === HOJA 2: DETALLE CHECK-OUTS ===
    ws_detalle = wb.create_sheet("🧾 Check-outs")
    headers_detalle = ["Fecha", "Habitación", "Huésped", "Total", "Cobrado", "Pendiente", "Método"]
    for c, h in enumerate(headers_detalle, 1):
        cell = ws_detalle.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    for r_idx, (fecha, hab, nom, ape, total, cobrado, pendiente, metodo) in enumerate(detalle_checkouts, 2):
        ws_detalle.cell(row=r_idx, column=1, value=fecha.strftime('%d/%m/%Y'))
        ws_detalle.cell(row=r_idx, column=2, value=hab)
        ws_detalle.cell(row=r_idx, column=3, value=f"{nom} {ape}")
        ws_detalle.cell(row=r_idx, column=4, value=total).number_format = '"S/ "#,##0.00'
        ws_detalle.cell(row=r_idx, column=5, value=cobrado).number_format = '"S/ "#,##0.00'
        ws_detalle.cell(row=r_idx, column=6, value=pendiente).number_format = '"S/ "#,##0.00'
        ws_detalle.cell(row=r_idx, column=7, value=metodo)
    
    # Fila de totales con FÓRMULAS
    last_row = len(detalle_checkouts) + 2
    ws_detalle.cell(row=last_row, column=1, value="TOTALES").font = Font(bold=True)
    ws_detalle.cell(row=last_row, column=4, value=f"=SUM(D2:D{last_row-1})").number_format = '"S/ "#,##0.00'
    ws_detalle.cell(row=last_row, column=5, value=f"=SUM(E2:E{last_row-1})").number_format = '"S/ "#,##0.00'
    ws_detalle.cell(row=last_row, column=6, value=f"=SUM(F2:F{last_row-1})").number_format = '"S/ "#,##0.00'
    
    for col in range(1, 8):
        ws_detalle.column_dimensions[get_column_letter(col)].width = 15

    # === HOJA 3: MINI-BAR ===
    ws_bar = wb.create_sheet("🍹 Mini-bar")
    headers_bar = ["Producto", "Categoría", "Unidades Vendidas", "Recaudación"]
    for c, h in enumerate(headers_bar, 1):
        cell = ws_bar.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    for r_idx, (prod, cat, cant, rec) in enumerate(consumo_bar, 2):
        ws_bar.cell(row=r_idx, column=1, value=prod)
        ws_bar.cell(row=r_idx, column=2, value=cat or '-')
        ws_bar.cell(row=r_idx, column=3, value=cant or 0)
        ws_bar.cell(row=r_idx, column=4, value=rec or 0).number_format = '"S/ "#,##0.00'
    
    last_bar = len(consumo_bar) + 2
    ws_bar.cell(row=last_bar, column=3, value=f"=SUM(C2:C{last_bar-1})")
    ws_bar.cell(row=last_bar, column=4, value=f"=SUM(D2:D{last_bar-1})").number_format = '"S/ "#,##0.00'
    
    for col in range(1, 5):
        ws_bar.column_dimensions[get_column_letter(col)].width = 20

    # === HOJA 4: DEUDAS PENDIENTES ===
    if deudas:
        ws_deudas = wb.create_sheet("⚠️ Deudas")
        headers_deuda = ["Habitación", "Huésped", "Monto Pendiente", "Fecha Salida"]
        for c, h in enumerate(headers_deuda, 1):
            cell = ws_deudas.cell(row=1, column=c, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="CC0000", end_color="CC0000", fill_type="solid")
        
        for r_idx, (hab, nom, ape, monto, fecha) in enumerate(deudas, 2):
            ws_deudas.cell(row=r_idx, column=1, value=hab)
            ws_deudas.cell(row=r_idx, column=2, value=f"{nom} {ape}")
            cell_monto = ws_deudas.cell(row=r_idx, column=3, value=monto)
            cell_monto.number_format = '"S/ "#,##0.00'
            cell_monto.font = Font(color="CC0000", bold=True)
            ws_deudas.cell(row=r_idx, column=4, value=fecha.strftime('%d/%m/%Y'))
        
        for col in range(1, 5):
            ws_deudas.column_dimensions[get_column_letter(col)].width = 20

    # Descargar
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, 
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, 
                     download_name=f"Reporte_Caja_{datetime.now().strftime('%Y%m%d')}.xlsx")
if __name__ == '__main__':
    inicializar_database()
    app.run(debug=True, host='0.0.0.0', port=8000)